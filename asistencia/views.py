from datetime import datetime, timedelta

from django.contrib import messages
from django.contrib.auth import login
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q
from django.http import JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST
from django.db import transaction
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .forms import (
    EjercicioForm,
    EntrenamientoInfoForm,
    JugadorForm,
    PerfilForm,
    RegistroEntrenadorForm,
    TrabajoTurnoForm,
    ObservacionJugadorForm,
    MotivoAusenciaForm,
    NoEntrenamientoForm,
    SetPartidoFormSet,
    PartidoTurnoForm,
)

from .models import (
    Asistencia,
    Ejercicio,
    EjercicioRealizado,
    EjercicioTurno,
    Entrenamiento,
    Entrenador,
    Jugador,
    TrabajoTurno,
    ObservacionJugador,
    PartidoTurno,
)
def obtener_o_crear_entrenamiento(fecha, turno):
    entrenamiento, _ = Entrenamiento.objects.get_or_create(fecha=fecha, turno=turno)
    return entrenamiento


def nombre_entrenador(entrenamiento):
    if entrenamiento.entrenador_responsable:
        return str(entrenamiento.entrenador_responsable)

    if entrenamiento.entrenador:
        return (
            entrenamiento.entrenador.get_full_name()
            or entrenamiento.entrenador.username
        )

    return "Sin entrenador"


def asignar_entrenador_si_vacio(entrenamiento, user):
    # Campo antiguo. Ya no se asigna automáticamente.
    # El entrenador real se elige desde la lista de entrenadores del turno.
    return


def turno_bloqueado(entrenamiento):
    return entrenamiento.finalizado or entrenamiento.no_se_entreno


def redirigir_turno_bloqueado(request, entrenamiento):
    messages.warning(
        request,
        "Este turno está bloqueado. Reabrilo o quitá la marca de no entrenamiento para modificarlo.",
    )

    return redirect_dia_turno(
        entrenamiento,
        request.POST.get("volver_a", ""),
    )


def redirect_dia_turno(entrenamiento, volver_a=""):
    url = (
        f"/dia/{entrenamiento.fecha.isoformat()}/"
        f"turno/{entrenamiento.turno}/"
    )

    if volver_a:
        url += volver_a

    return redirect(url)


def registro(request):
    if request.method == "POST":
        form = RegistroEntrenadorForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request, "Cuenta creada correctamente.")
            return redirect("inicio")
    else:
        form = RegistroEntrenadorForm()

    return render(request, "registration/register.html", {"form": form})


@login_required
def inicio(request):
    fecha_str = request.GET.get("fecha")

    if fecha_str:
        try:
            fecha = datetime.strptime(fecha_str, "%Y-%m-%d").date()
        except ValueError:
            fecha = timezone.localdate()
    else:
        fecha = timezone.localdate()

    turnos_info = []
    total_cargados_dia = 0
    total_marcados_dia = 0
    total_no_entrenados_dia = 0
    total_finalizados_dia = 0

    for turno in [1, 2, 3]:
        entrenamiento = obtener_o_crear_entrenamiento(fecha, turno)

        cantidad_jugadores = Asistencia.objects.filter(
            entrenamiento=entrenamiento
        ).count()

        cantidad_marcados = (
            Asistencia.objects
            .filter(entrenamiento=entrenamiento)
            .exclude(estado="pendiente")
            .count()
        )

        if not entrenamiento.no_se_entreno:
            total_cargados_dia += cantidad_jugadores
            total_marcados_dia += cantidad_marcados

        if entrenamiento.no_se_entreno:
            total_no_entrenados_dia += 1

        if entrenamiento.finalizado:
            total_finalizados_dia += 1

        if entrenamiento.no_se_entreno:
            estado_texto = "No se entrenó"
            estado_clase = "danger"
            estado_detalle = entrenamiento.get_motivo_no_entrenamiento_display()

            if entrenamiento.detalle_no_entrenamiento:
                estado_detalle = (
                    f"{estado_detalle} — "
                    f"{entrenamiento.detalle_no_entrenamiento}"
                )

        elif entrenamiento.finalizado:
            estado_texto = "Finalizado"
            estado_clase = "success"
            estado_detalle = "Turno cerrado correctamente."

        elif cantidad_jugadores == 0:
            estado_texto = "Sin cargar"
            estado_clase = "secondary"
            estado_detalle = "Todavía no hay jugadores cargados."

        elif cantidad_marcados < cantidad_jugadores:
            estado_texto = "Datos pendientes"
            estado_clase = "warning"
            estado_detalle = (
                f"{cantidad_marcados} de "
                f"{cantidad_jugadores} asistencias marcadas."
            )

        else:
            estado_texto = "Asistencias marcadas"
            estado_clase = "primary"
            estado_detalle = "Falta revisar cierre del turno."

        turnos_info.append({
            "turno": turno,
            "cantidad_jugadores": cantidad_jugadores,
            "cantidad_marcados": cantidad_marcados,
            "entrenador": nombre_entrenador(entrenamiento),
            "observaciones": entrenamiento.observaciones,
            "no_se_entreno": entrenamiento.no_se_entreno,
            "motivo_no_entrenamiento": (
                entrenamiento.get_motivo_no_entrenamiento_display()
                if entrenamiento.no_se_entreno
                else ""
            ),
            "detalle_no_entrenamiento": entrenamiento.detalle_no_entrenamiento,
            "finalizado": entrenamiento.finalizado,
            "estado_texto": estado_texto,
            "estado_clase": estado_clase,
            "estado_detalle": estado_detalle,
        })

    contexto = {
        "fecha": fecha,
        "hoy": timezone.localdate(),
        "ayer": fecha - timedelta(days=1),
        "maniana": fecha + timedelta(days=1),
        "turnos_info": turnos_info,
        "total_jugadores": Jugador.objects.filter(activo=True).count(),
        "total_ejercicios": Ejercicio.objects.filter(activo=True).count(),
        "total_cargados_dia": total_cargados_dia,
        "total_marcados_dia": total_marcados_dia,
        "total_no_entrenados_dia": total_no_entrenados_dia,
        "total_finalizados_dia": total_finalizados_dia,
    }

    return render(request, "asistencia/inicio.html", contexto)


@login_required
def ir_a_fecha_asistencia(request):
    fecha_str = request.GET.get("fecha")
    turno = request.GET.get("turno", 1)

    if not fecha_str:
        fecha_str = timezone.localdate().isoformat()

    return redirect("dia_turno", fecha_str=fecha_str, turno=int(turno))


@login_required
def dia_turno(request, fecha_str, turno):
    fecha = datetime.strptime(
        fecha_str,
        "%Y-%m-%d",
    ).date()

    turno = int(turno)

    entrenamiento = obtener_o_crear_entrenamiento(
        fecha,
        turno,
    )

    asistencias = list(
        Asistencia.objects
        .filter(entrenamiento=entrenamiento)
        .select_related("jugador")
    )

    jugadores_disponibles = (
        Jugador.objects
        .filter(activo=True)
        .exclude(
            id__in=[
                asistencia.jugador_id
                for asistencia in asistencias
            ]
        )
        .order_by(
            "apellido",
            "nombre",
        )
    )

    ejercicios_turno = (
        EjercicioTurno.objects
        .filter(entrenamiento=entrenamiento)
        .select_related("ejercicio")
        .order_by(
            "ejercicio__categoria",
            "ejercicio__nombre",
        )
    )

    ejercicios_turno_por_categoria = {}

    for ejercicio_turno in ejercicios_turno:
        categoria = ejercicio_turno.ejercicio.get_categoria_display()

        ejercicios_turno_por_categoria.setdefault(
            categoria,
            [],
        ).append(
            ejercicio_turno.ejercicio.nombre
        )

    total_ejercicios_turno = ejercicios_turno.count()

    for asistencia in asistencias:
        asistencia.observaciones_turno = (
            ObservacionJugador.objects
            .filter(
                jugador=asistencia.jugador,
                entrenamiento=entrenamiento,
            )
            .select_related("creada_por")
            .order_by("-creada_el")
        )

    trabajos_turno = (
        TrabajoTurno.objects
        .filter(entrenamiento=entrenamiento)
        .select_related(
            "jugador_1",
            "jugador_2",
        )
        .order_by(
            "cambio",
            "id",
        )
    )

    jugadores_ocupados_por_cambio = {}

    for trabajo in trabajos_turno:
        cambio_clave = str(trabajo.cambio)

        if cambio_clave not in jugadores_ocupados_por_cambio:
            jugadores_ocupados_por_cambio[cambio_clave] = []

        jugadores_ocupados_por_cambio[cambio_clave].append(
            trabajo.jugador_1_id
        )

        if trabajo.jugador_2_id:
            jugadores_ocupados_por_cambio[cambio_clave].append(
                trabajo.jugador_2_id
            )

    cambios_resumen = []

    numeros_cambio = sorted(
        set(
            trabajos_turno.values_list(
                "cambio",
                flat=True,
            )
        )
    )

    for numero_cambio in numeros_cambio:
        trabajos_del_cambio = trabajos_turno.filter(
            cambio=numero_cambio
        )

        jugadores_asignados_ids = set()

        for trabajo in trabajos_del_cambio:
            jugadores_asignados_ids.add(
                trabajo.jugador_1_id
            )

            if trabajo.jugador_2_id:
                jugadores_asignados_ids.add(
                    trabajo.jugador_2_id
                )

        jugadores_pendientes = [
            asistencia.jugador
            for asistencia in asistencias
            if asistencia.jugador_id not in jugadores_asignados_ids
        ]

        cantidad_asignados = len(jugadores_asignados_ids)
        cantidad_total = len(asistencias)

        cambios_resumen.append({
            "numero": numero_cambio,
            "cantidad_asignados": cantidad_asignados,
            "cantidad_total": cantidad_total,
            "jugadores_pendientes": jugadores_pendientes,
            "completo": (
                cantidad_total > 0
                and cantidad_asignados == cantidad_total
            ),
        })

    trabajos_otros_turnos = (
        TrabajoTurno.objects
        .filter(
            entrenamiento__fecha=fecha,
            entrenamiento__turno__lt=turno,
        )
        .select_related(
            "entrenamiento",
            "jugador_1",
            "jugador_2",
        )
        .order_by(
            "entrenamiento__turno",
            "cambio",
            "id",
        )
    )

    partidos_turno = (
        PartidoTurno.objects
        .filter(entrenamiento=entrenamiento)
        .select_related(
            "jugador_1",
            "jugador_2",
        )
        .prefetch_related("sets")
        .order_by("id")
    )

    total_jugadores_turno = len(asistencias)

    total_presentes_turno = sum(
        1
        for asistencia in asistencias
        if asistencia.estado == "asistio"
    )

    total_tardes_turno = sum(
        1
        for asistencia in asistencias
        if asistencia.estado == "tarde"
    )

    total_ausentes_turno = sum(
        1
        for asistencia in asistencias
        if asistencia.estado == "ausente"
    )

    total_pendientes_turno = sum(
        1
        for asistencia in asistencias
        if asistencia.estado == "pendiente"
    )

    ausentes_sin_motivo_turno = sum(
        1
        for asistencia in asistencias
        if (
            asistencia.estado == "ausente"
            and not asistencia.motivo_ausencia
        )
    )

    total_trabajos_turno = trabajos_turno.count()
    total_cambios_turno = len(cambios_resumen)

    cambios_completos_turno = sum(
        1
        for cambio in cambios_resumen
        if cambio["completo"]
    )

    cambios_incompletos_turno = (
        total_cambios_turno
        - cambios_completos_turno
    )

    total_partidos_turno = partidos_turno.count()

    partidos_sin_sets_turno = sum(
        1
        for partido in partidos_turno
        if not partido.sets.all()
    )

    total_observaciones_turno = (
        ObservacionJugador.objects
        .filter(entrenamiento=entrenamiento)
        .count()
    )

    alertas_resumen_turno = []

    if total_jugadores_turno == 0:
        alertas_resumen_turno.append(
            "No hay jugadores cargados."
        )

    if total_pendientes_turno > 0:
        alertas_resumen_turno.append(
            (
                f"Hay {total_pendientes_turno} "
                "jugador(es) con asistencia pendiente."
            )
        )

    if ausentes_sin_motivo_turno > 0:
        alertas_resumen_turno.append(
            (
                f"Hay {ausentes_sin_motivo_turno} "
                "ausencia(s) sin motivo."
            )
        )

    if cambios_incompletos_turno > 0:
        alertas_resumen_turno.append(
            (
                f"Hay {cambios_incompletos_turno} "
                "cambio(s) incompleto(s)."
            )
        )

    if partidos_sin_sets_turno > 0:
        alertas_resumen_turno.append(
            (
                f"Hay {partidos_sin_sets_turno} "
                "partido(s) sin sets."
            )
        )

    if entrenamiento.entrenador_responsable is None:
        alertas_resumen_turno.append(
            "El turno no tiene entrenador responsable."
        )

    resumen_turno = {
        "total_jugadores": total_jugadores_turno,
        "presentes": total_presentes_turno,
        "tardes": total_tardes_turno,
        "ausentes": total_ausentes_turno,
        "pendientes": total_pendientes_turno,
        "ausentes_sin_motivo": ausentes_sin_motivo_turno,
        "total_trabajos": total_trabajos_turno,
        "total_cambios": total_cambios_turno,
        "cambios_completos": cambios_completos_turno,
        "cambios_incompletos": cambios_incompletos_turno,
        "total_partidos": total_partidos_turno,
        "partidos_sin_sets": partidos_sin_sets_turno,
        "total_observaciones": total_observaciones_turno,
        "total_ejercicios": total_ejercicios_turno,
        "alertas": alertas_resumen_turno,
        "listo_para_finalizar": (
            total_jugadores_turno > 0
            and total_pendientes_turno == 0
            and ausentes_sin_motivo_turno == 0
            and cambios_incompletos_turno == 0
            and partidos_sin_sets_turno == 0
            and entrenamiento.entrenador_responsable is not None 
            ),
    }

    contexto = {
        "entrenamiento": entrenamiento,
        "entrenamiento_form": EntrenamientoInfoForm(
            instance=entrenamiento
        ),
        "no_entrenamiento_form": NoEntrenamientoForm(
            instance=entrenamiento
        ),
        "trabajo_form": TrabajoTurnoForm(
            entrenamiento=entrenamiento
        ),
        "ejercicios_turno": ejercicios_turno,
        "ejercicios_turno_por_categoria": ejercicios_turno_por_categoria,
        "trabajos_turno": trabajos_turno,
        "jugadores_ocupados_por_cambio": jugadores_ocupados_por_cambio,
        "cambios_resumen": cambios_resumen,
        "trabajos_otros_turnos": trabajos_otros_turnos,
        "nombre_entrenador": nombre_entrenador(
            entrenamiento
        ),  
        "asistencias": asistencias,
        "jugadores_disponibles": jugadores_disponibles,
        "partidos_turno": partidos_turno,
        "resumen_turno": resumen_turno,
        "fecha": fecha,
        "turno": turno,
        "ayer": fecha - timedelta(days=1),
        "maniana": fecha + timedelta(days=1),
        "hoy": timezone.localdate(),
    }

    return render(
        request,
        "asistencia/dia_turno.html",
        contexto,
    )
    
@login_required
def editar_trabajo_turno(request, trabajo_id):
    trabajo = get_object_or_404(
        TrabajoTurno,
        id=trabajo_id,
    )

    entrenamiento = trabajo.entrenamiento

    if request.method == "POST" and turno_bloqueado(entrenamiento):
        return redirigir_turno_bloqueado(
            request,
            entrenamiento,
        )

    if request.method == "POST":
        form = TrabajoTurnoForm(
            request.POST,
            instance=trabajo,
            entrenamiento=entrenamiento,
        )

        if form.is_valid():
            form.save()

            messages.success(
                request,
                "Trabajo actualizado correctamente.",
            )

            return redirect(
                "dia_turno",
                fecha_str=entrenamiento.fecha.isoformat(),
                turno=entrenamiento.turno,
            )
    else:
        form = TrabajoTurnoForm(
            instance=trabajo,
            entrenamiento=entrenamiento,
        )

    return render(
        request,
        "asistencia/editar_trabajo_turno.html",
        {
            "form": form,
            "trabajo": trabajo,
            "entrenamiento": entrenamiento,
        },
    )



@login_required
@require_POST
def agregar_trabajo_turno(request, entrenamiento_id):
    entrenamiento = get_object_or_404(
        Entrenamiento,
        id=entrenamiento_id,
    )

    if turno_bloqueado(entrenamiento):
        return redirigir_turno_bloqueado(
            request,
            entrenamiento,
        )

    form = TrabajoTurnoForm(
        request.POST,
        entrenamiento=entrenamiento,
    )

    if form.is_valid():
        trabajo = form.save(commit=False)
        trabajo.entrenamiento = entrenamiento
        trabajo.save()

        messages.success(
            request,
            f"Trabajo agregado correctamente al cambio {trabajo.cambio}.",
        )
    else:
        errores = []

        for lista_errores in form.errors.values():
            for error in lista_errores:
                errores.append(str(error))

        if errores:
            messages.error(
                request,
                "No se pudo agregar el trabajo: " + " ".join(errores),
            )
        else:
            messages.error(
                request,
                "No se pudo agregar el trabajo. Revisá los datos ingresados.",
            )

    return redirect_dia_turno(
        entrenamiento,
        request.POST.get("volver_a", ""),
    )
    
@login_required
def crear_partido_turno(request, entrenamiento_id):
    entrenamiento = get_object_or_404(
        Entrenamiento,
        id=entrenamiento_id,
    )

    if turno_bloqueado(entrenamiento):
        return redirigir_turno_bloqueado(
            request,
            entrenamiento,
        )

    if request.method == "POST":
        partido_form = PartidoTurnoForm(
            request.POST,
            entrenamiento=entrenamiento,
        )

        partido_temporal = PartidoTurno(
            entrenamiento=entrenamiento
        )

        sets_formset = SetPartidoFormSet(
            request.POST,
            instance=partido_temporal,
            prefix="sets",
        )

        if partido_form.is_valid() and sets_formset.is_valid():
            with transaction.atomic():
                partido = partido_form.save(commit=False)
                partido.entrenamiento = entrenamiento
                partido.save()

                sets_formset.instance = partido

                sets_guardados = sets_formset.save(
                    commit=False
                )

                numero_set = 1

                for set_partido in sets_guardados:
                    set_partido.partido = partido
                    set_partido.numero = numero_set
                    set_partido.save()

                    numero_set += 1

                for set_eliminado in sets_formset.deleted_objects:
                    if set_eliminado.pk:
                        set_eliminado.delete()

            messages.success(
                request,
                (
                    f"Partido guardado: "
                    f"{partido.jugador_1} "
                    f"{partido.resultado_general} "
                    f"{partido.jugador_2}."
                ),
            )

            return redirect(
                "dia_turno",
                fecha_str=entrenamiento.fecha.isoformat(),
                turno=entrenamiento.turno,
            )
    else:
        partido_form = PartidoTurnoForm(
            entrenamiento=entrenamiento,
        )

        partido_temporal = PartidoTurno(
            entrenamiento=entrenamiento
        )

        sets_formset = SetPartidoFormSet(
            instance=partido_temporal,
            prefix="sets",
        )

    return render(
        request,
        "asistencia/crear_partido_turno.html",
        {
            "entrenamiento": entrenamiento,
            "partido_form": partido_form,
            "sets_formset": sets_formset,
        },
    )


@login_required
@require_POST
def eliminar_trabajo_turno(request, trabajo_id):
    trabajo = get_object_or_404(
        TrabajoTurno,
        id=trabajo_id,
    )

    entrenamiento = trabajo.entrenamiento
    numero_cambio = trabajo.cambio

    if turno_bloqueado(entrenamiento):
        return redirigir_turno_bloqueado(
            request,
            entrenamiento,
        )

    trabajo.delete()

    messages.success(
        request,
        f"Trabajo del cambio {numero_cambio} eliminado correctamente.",
    )

    return redirect_dia_turno(
        entrenamiento,
        request.POST.get("volver_a", ""),
    )


@login_required
@require_POST
def tomar_turno(request, entrenamiento_id):
    entrenamiento = get_object_or_404(
        Entrenamiento,
        id=entrenamiento_id,
    )

    if entrenamiento.entrenador is None:
        entrenamiento.entrenador = request.user
        entrenamiento.save()
        messages.success(
            request,
            "Tomaste este turno correctamente.",
        )

    elif entrenamiento.entrenador == request.user:
        messages.info(
            request,
            "Este turno ya está asignado a vos.",
        )

    else:
        messages.warning(
            request,
            "Este turno ya fue tomado por otro entrenador.",
        )

    return redirect(
        "dia_turno",
        fecha_str=entrenamiento.fecha.isoformat(),
        turno=entrenamiento.turno,
    )


@login_required
@require_POST
def guardar_info_entrenamiento(request, entrenamiento_id):
    entrenamiento = get_object_or_404(
        Entrenamiento,
        id=entrenamiento_id,
    )

    if turno_bloqueado(entrenamiento):
        return redirigir_turno_bloqueado(
            request,
            entrenamiento,
        )

    form = EntrenamientoInfoForm(
        request.POST,
        instance=entrenamiento,
    )

    if form.is_valid():
        form.save()

        messages.success(
            request,
            "Información del turno guardada correctamente.",
        )
    else:
        messages.error(
            request,
            "No se pudo guardar la información del turno. Revisá los datos.",
        )

    return redirect_dia_turno(
        entrenamiento,
        request.POST.get("volver_a", ""),
    )
    
@login_required
@require_POST
def guardar_no_entrenamiento(request, entrenamiento_id):
    entrenamiento = get_object_or_404(
        Entrenamiento,
        id=entrenamiento_id,
    )

    if entrenamiento.finalizado:
        messages.warning(
            request,
            "Este turno está finalizado. Reabrilo para modificarlo.",
        )

        return redirect_dia_turno(
            entrenamiento,
            request.POST.get("volver_a", ""),
        )

    accion = request.POST.get("accion")

    if accion == "quitar":
        entrenamiento.no_se_entreno = False
        entrenamiento.motivo_no_entrenamiento = ""
        entrenamiento.detalle_no_entrenamiento = ""

        entrenamiento.save(
            update_fields=[
                "no_se_entreno",
                "motivo_no_entrenamiento",
                "detalle_no_entrenamiento",
            ]
        )

        messages.success(
            request,
            "Se quitó la marca de no entrenamiento. El turno quedó abierto para cargar datos.",
        )

        return redirect_dia_turno(
            entrenamiento,
            request.POST.get("volver_a", ""),
        )

    form = NoEntrenamientoForm(
        request.POST,
        instance=entrenamiento,
    )

    if form.is_valid():
        Asistencia.objects.filter(
            entrenamiento=entrenamiento,
        ).delete()

        TrabajoTurno.objects.filter(
            entrenamiento=entrenamiento,
        ).delete()

        PartidoTurno.objects.filter(
            entrenamiento=entrenamiento,
        ).delete()

        EjercicioTurno.objects.filter(
            entrenamiento=entrenamiento,
        ).delete()

        ObservacionJugador.objects.filter(
            entrenamiento=entrenamiento,
        ).delete()

        turno_sin_entrenamiento = form.save(commit=False)
        turno_sin_entrenamiento.no_se_entreno = True
        turno_sin_entrenamiento.finalizado = False
        turno_sin_entrenamiento.finalizado_el = None
        turno_sin_entrenamiento.finalizado_por = None
        turno_sin_entrenamiento.save()

        messages.success(
            request,
            "El turno fue marcado como no entrenado y quedó bloqueado.",
        )
    else:
        messages.error(
            request,
            "No se pudo guardar. Seleccioná un motivo.",
        )

    return redirect_dia_turno(
        entrenamiento,
        request.POST.get("volver_a", ""),
    )


@login_required
@require_POST
def finalizar_turno(request, entrenamiento_id):
    entrenamiento = get_object_or_404(
        Entrenamiento,
        id=entrenamiento_id,
    )

    if entrenamiento.finalizado:
        messages.info(
            request,
            "Este turno ya está finalizado.",
        )

        return redirect(
            "dia_turno",
            fecha_str=entrenamiento.fecha.isoformat(),
            turno=entrenamiento.turno,
        )

    if entrenamiento.no_se_entreno:
        messages.warning(
            request,
            "Este turno está marcado como no entrenado. Quitá esa marca si querés cargar datos o finalizarlo como entrenamiento.",
        )

        return redirect(
            "dia_turno",
            fecha_str=entrenamiento.fecha.isoformat(),
            turno=entrenamiento.turno,
        )

    errores = []

    asistencias = (
        Asistencia.objects
        .filter(entrenamiento=entrenamiento)
        .select_related("jugador")
    )

    if not asistencias.exists():
        errores.append(
            "El turno no tiene jugadores cargados."
        )

    jugadores_pendientes = [
        str(asistencia.jugador)
        for asistencia in asistencias
        if asistencia.estado == "pendiente"
    ]

    if jugadores_pendientes:
        errores.append(
            "Falta marcar la asistencia de: "
            + ", ".join(jugadores_pendientes)
            + "."
        )

    ausentes_sin_motivo = [
        str(asistencia.jugador)
        for asistencia in asistencias
        if (
            asistencia.estado == "ausente"
            and not asistencia.motivo_ausencia
        )
    ]

    if ausentes_sin_motivo:
        errores.append(
            "Falta cargar el motivo de ausencia de: "
            + ", ".join(ausentes_sin_motivo)
            + "."
        )

    if entrenamiento.entrenador_responsable is None:
        errores.append(
            "El turno no tiene un entrenador responsable."
        )

    jugadores_que_entrenaron_ids = set(
        asistencias
        .filter(
            Q(estado="asistio")
            | Q(estado="tarde")
        )
        .values_list(
            "jugador_id",
            flat=True,
        )
    )

    trabajos = (
        TrabajoTurno.objects
        .filter(entrenamiento=entrenamiento)
        .select_related(
            "jugador_1",
            "jugador_2",
        )
    )

    numeros_cambio = sorted(
        set(
            trabajos.values_list(
                "cambio",
                flat=True,
            )
        )
    )

    for numero_cambio in numeros_cambio:
        trabajos_del_cambio = trabajos.filter(
            cambio=numero_cambio
        )

        jugadores_asignados_ids = set()

        for trabajo in trabajos_del_cambio:
            jugadores_asignados_ids.add(
                trabajo.jugador_1_id
            )

            if trabajo.jugador_2_id:
                jugadores_asignados_ids.add(
                    trabajo.jugador_2_id
                )

        jugadores_faltantes_ids = (
            jugadores_que_entrenaron_ids
            - jugadores_asignados_ids
        )

        if jugadores_faltantes_ids:
            jugadores_faltantes = (
                Jugador.objects
                .filter(id__in=jugadores_faltantes_ids)
                .order_by(
                    "apellido",
                    "nombre",
                )
            )

            errores.append(
                f"El cambio {numero_cambio} está incompleto. "
                f"Faltan: "
                + ", ".join(
                    str(jugador)
                    for jugador in jugadores_faltantes
                )
                + "."
            )

    partidos_sin_sets = (
        PartidoTurno.objects
        .filter(entrenamiento=entrenamiento)
        .annotate(cantidad_sets=Count("sets"))
        .filter(cantidad_sets=0)
        .select_related(
            "jugador_1",
            "jugador_2",
        )
    )

    for partido in partidos_sin_sets:
        errores.append(
            f"El partido {partido.jugador_1} vs "
            f"{partido.jugador_2} no tiene sets cargados."
        )

    if errores:
        for error in errores:
            messages.error(
                request,
                error,
            )

        messages.warning(
            request,
            "El turno no pudo finalizarse. Revisá los datos indicados.",
        )

        return redirect(
            "dia_turno",
            fecha_str=entrenamiento.fecha.isoformat(),
            turno=entrenamiento.turno,
        )

    entrenamiento.finalizado = True
    entrenamiento.finalizado_el = timezone.now()
    entrenamiento.finalizado_por = request.user

    entrenamiento.save(
        update_fields=[
            "finalizado",
            "finalizado_el",
            "finalizado_por",
        ]
    )

    messages.success(
        request,
        "Turno finalizado correctamente.",
    )

    return redirect(
        "dia_turno",
        fecha_str=entrenamiento.fecha.isoformat(),
        turno=entrenamiento.turno,
    )



@login_required
@require_POST
def reabrir_turno(request, entrenamiento_id):
    entrenamiento = get_object_or_404(
        Entrenamiento,
        id=entrenamiento_id,
    )

    if not entrenamiento.finalizado:
        messages.info(
            request,
            "Este turno ya está abierto.",
        )

        return redirect(
            "dia_turno",
            fecha_str=entrenamiento.fecha.isoformat(),
            turno=entrenamiento.turno,
        )

    entrenamiento.finalizado = False
    entrenamiento.finalizado_el = None
    entrenamiento.finalizado_por = None

    entrenamiento.save(
        update_fields=[
            "finalizado",
            "finalizado_el",
            "finalizado_por",
        ]
    )

    messages.success(
        request,
        "El turno fue reabierto correctamente.",
    )

    return redirect(
        "dia_turno",
        fecha_str=entrenamiento.fecha.isoformat(),
        turno=entrenamiento.turno,
    )

@login_required
@require_POST
def guardar_observacion_jugador(request, asistencia_id):
    asistencia = get_object_or_404(
        Asistencia.objects.select_related(
            "jugador",
            "entrenamiento",
        ),
        id=asistencia_id,
    )

    entrenamiento = asistencia.entrenamiento
    jugador = asistencia.jugador

    if turno_bloqueado(entrenamiento):
        return redirigir_turno_bloqueado(
            request,
            entrenamiento,
        )

    form = ObservacionJugadorForm(request.POST)

    if form.is_valid():
        observacion = form.save(commit=False)
        observacion.jugador = jugador
        observacion.entrenamiento = entrenamiento
        observacion.creada_por = request.user
        observacion.save()

        messages.success(
            request,
            f"Observación guardada para {jugador}.",
        )
    else:
        errores = []

        for lista_errores in form.errors.values():
            for error in lista_errores:
                errores.append(str(error))

        messages.error(
            request,
            (
                "No se pudo guardar la observación: "
                + " ".join(errores)
            ),
        )

    return redirect_dia_turno(
        entrenamiento,
        request.POST.get("volver_a", ""),
    )




@login_required
@require_POST
def agregar_jugador(request, entrenamiento_id):
    entrenamiento = get_object_or_404(
        Entrenamiento,
        id=entrenamiento_id,
    )

    if turno_bloqueado(entrenamiento):
        return redirigir_turno_bloqueado(
            request,
            entrenamiento,
        )

    jugadores_ids = request.POST.getlist("jugadores_ids")

    if not jugadores_ids:
        messages.info(
            request,
            "Seleccioná al menos un jugador para agregar al turno.",
        )

        return redirect(
            "dia_turno",
            fecha_str=entrenamiento.fecha.strftime("%Y-%m-%d"),
            turno=entrenamiento.turno,
        )

    jugadores = (
        Jugador.objects
        .filter(
            id__in=jugadores_ids,
            activo=True,
        )
        .order_by(
            "apellido",
            "nombre",
        )
    )

    agregados = 0
    ya_cargados = 0

    for jugador in jugadores:
        asistencia, creado = Asistencia.objects.get_or_create(
            entrenamiento=entrenamiento,
            jugador=jugador,
            defaults={
                "estado": "pendiente",
            },
        )

        if creado:
            agregados += 1
        else:
            ya_cargados += 1

    if agregados > 0:
        messages.success(
            request,
            f"Se agregaron {agregados} jugador(es) al Turno {entrenamiento.turno}.",
        )

    if ya_cargados > 0:
        messages.info(
            request,
            f"{ya_cargados} jugador(es) ya estaban cargados en este turno.",
        )

    return redirect_dia_turno(
        entrenamiento,
        request.POST.get("volver_a", ""),
    )


@login_required
@require_POST
def copiar_lista_ayer(request, entrenamiento_id):
    entrenamiento = get_object_or_404(Entrenamiento, id=entrenamiento_id)

    if turno_bloqueado(entrenamiento):
        return redirigir_turno_bloqueado(
            request,
            entrenamiento,
        )

    entrenamiento_anterior = (
        Entrenamiento.objects
        .filter(
            fecha__lt=entrenamiento.fecha,
            turno=entrenamiento.turno,
            asistencias__isnull=False
        )
        .annotate(total_jugadores=Count("asistencias"))
        .filter(total_jugadores__gt=0)
        .order_by("-fecha")
        .first()
    )

    if not entrenamiento_anterior:
        messages.info(request, "No se encontró una lista anterior para copiar.")
        return redirect("dia_turno", fecha_str=entrenamiento.fecha.strftime("%Y-%m-%d"), turno=entrenamiento.turno)

    asistencias_anteriores = Asistencia.objects.filter(
        entrenamiento=entrenamiento_anterior
    ).select_related("jugador")

    jugadores_copiados = 0

    for asistencia_anterior in asistencias_anteriores:
        _, creado = Asistencia.objects.get_or_create(
            entrenamiento=entrenamiento,
            jugador=asistencia_anterior.jugador,
            defaults={"estado": "pendiente"}
        )

        if creado:
            jugadores_copiados += 1

    if jugadores_copiados > 0:
        messages.success(
            request,
            f"Lista anterior copiada correctamente. Se agregaron {jugadores_copiados} jugador/es."
        )
    else:
        messages.info(
            request,
            "La lista anterior ya estaba cargada en este turno."
        )

    return redirect("dia_turno", fecha_str=entrenamiento.fecha.strftime("%Y-%m-%d"), turno=entrenamiento.turno)


@login_required
@require_POST
def marcar_todos_asistieron(request, entrenamiento_id):
    entrenamiento = get_object_or_404(Entrenamiento, id=entrenamiento_id)
    if turno_bloqueado(entrenamiento):
        return redirigir_turno_bloqueado(
            request,
            entrenamiento,
        )

    asignar_entrenador_si_vacio(entrenamiento, request.user)

    Asistencia.objects.filter(entrenamiento=entrenamiento).update(estado="asistio")

    messages.success(request, "Todos los jugadores quedaron como asistieron.")
    return redirect_dia_turno(
        entrenamiento,
        request.POST.get("volver_a", ""),
    )


@login_required
@require_POST
def quitar_jugador(request, asistencia_id):
    asistencia = get_object_or_404(
        Asistencia,
        id=asistencia_id,
    )

    entrenamiento = asistencia.entrenamiento
    jugador = asistencia.jugador

    if turno_bloqueado(entrenamiento):
        return redirigir_turno_bloqueado(
            request,
            entrenamiento,
        )

    asignar_entrenador_si_vacio(
        entrenamiento,
        request.user,
    )

    trabajos_eliminados, _ = (
        TrabajoTurno.objects
        .filter(entrenamiento=entrenamiento)
        .filter(
            Q(jugador_1=jugador)
            | Q(jugador_2=jugador)
        )
        .delete()
    )

    asistencia.delete()

    if trabajos_eliminados > 0:
        messages.success(
            request,
            (
                f"{jugador} fue quitado del turno y también se eliminaron "
                f"sus trabajos o parejas cargadas."
            ),
        )
    else:
        messages.success(
            request,
            f"{jugador} fue quitado del turno.",
        )

    return redirect_dia_turno(
        entrenamiento,
        request.POST.get("volver_a", ""),
    )


@login_required
@require_POST
def cambiar_estado(request, asistencia_id):
    asistencia = get_object_or_404(
        Asistencia.objects.select_related("entrenamiento"),
        id=asistencia_id,
    )

    if turno_bloqueado(asistencia.entrenamiento):
        return JsonResponse({
            "ok": False,
            "error": "Este turno está bloqueado.",
        }, status=403)

    asignar_entrenador_si_vacio(asistencia.entrenamiento, request.user)

    estado_nuevo = request.POST.get("estado")
    estados_validos = {"asistio", "ausente", "tarde"}

    if asistencia.estado == estado_nuevo:
      asistencia.estado = "pendiente"
    else:
      asistencia.estado = estado_nuevo

    if asistencia.estado != "ausente":
      asistencia.motivo_ausencia = ""
    asistencia.detalle_ausencia = ""

    asistencia.save()

    return JsonResponse({
        "ok": True,
        "estado": asistencia.estado,
    })
    
@login_required
@require_POST
def guardar_motivo_ausencia(request, asistencia_id):
    asistencia = get_object_or_404(
        Asistencia.objects.select_related(
            "jugador",
            "entrenamiento",
        ),
        id=asistencia_id,
    )

    if turno_bloqueado(asistencia.entrenamiento):
        return redirigir_turno_bloqueado(
            request,
            asistencia.entrenamiento,
        )

    form = MotivoAusenciaForm(
        request.POST,
        instance=asistencia,
    )

    if asistencia.estado != "ausente":
        messages.error(
            request,
            "Solo podés cargar un motivo cuando el jugador está ausente.",
        )
    elif form.is_valid():
        form.save()

        messages.success(
            request,
            f"Motivo de ausencia guardado para {asistencia.jugador}.",
        )
    else:
        errores = []

        for lista_errores in form.errors.values():
            for error in lista_errores:
                errores.append(str(error))

        messages.error(
            request,
            "No se pudo guardar el motivo: " + " ".join(errores),
        )

    return redirect_dia_turno(
        entrenamiento,
        request.POST.get("volver_a", ""),
    )


@login_required
def lista_jugadores(request):
    jugadores = Jugador.objects.all()
    return render(request, "asistencia/lista_jugadores.html", {"jugadores": jugadores})


@login_required
def crear_jugador(request):
    if request.method == "POST":
        form = JugadorForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Jugador agregado correctamente.")
            return redirect("lista_jugadores")
    else:
        form = JugadorForm()

    return render(request, "asistencia/form_jugador.html", {
        "form": form,
        "titulo": "Agregar jugador",
    })


@login_required
def editar_jugador(request, pk):
    jugador = get_object_or_404(Jugador, pk=pk)

    if request.method == "POST":
        form = JugadorForm(request.POST, instance=jugador)
        if form.is_valid():
            form.save()
            messages.success(request, "Jugador editado correctamente.")
            return redirect("lista_jugadores")
    else:
        form = JugadorForm(instance=jugador)

    return render(request, "asistencia/form_jugador.html", {
        "form": form,
        "titulo": "Editar jugador",
    })


@login_required
def lista_ejercicios(request):
    ejercicios = Ejercicio.objects.all()
    return render(request, "asistencia/lista_ejercicios.html", {"ejercicios": ejercicios})


@login_required
def crear_ejercicio(request):
    if request.method == "POST":
        form = EjercicioForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Ejercicio agregado correctamente.")
            return redirect("lista_ejercicios")
    else:
        form = EjercicioForm()

    return render(request, "asistencia/form_ejercicio.html", {
        "form": form,
        "titulo": "Agregar ejercicio",
    })


@login_required
def editar_ejercicio(request, pk):
    ejercicio = get_object_or_404(Ejercicio, pk=pk)

    if request.method == "POST":
        form = EjercicioForm(request.POST, instance=ejercicio)
        if form.is_valid():
            form.save()
            messages.success(request, "Ejercicio editado correctamente.")
            return redirect("lista_ejercicios")
    else:
        form = EjercicioForm(instance=ejercicio)

    return render(request, "asistencia/form_ejercicio.html", {
        "form": form,
        "titulo": "Editar ejercicio",
    })


@login_required
def cargar_ejercicios(request):
    fecha_str = request.GET.get("fecha")
    turno_str = request.GET.get("turno", 1)

    if fecha_str:
        try:
            fecha = datetime.strptime(
                fecha_str,
                "%Y-%m-%d",
            ).date()
        except ValueError:
            fecha = timezone.localdate()
    else:
        fecha = timezone.localdate()

    try:
        turno = int(turno_str)
    except (TypeError, ValueError):
        turno = 1

    if turno not in [1, 2, 3]:
        turno = 1

    entrenamiento = obtener_o_crear_entrenamiento(
        fecha,
        turno,
    )

    ejercicios_guardados = list(
        EjercicioTurno.objects
        .filter(entrenamiento=entrenamiento)
        .values_list("ejercicio_id", flat=True)
    )

    ejercicios_por_categoria = {
        "Movilidad": Ejercicio.objects.filter(
            categoria=Ejercicio.Categoria.MOVILIDAD,
            activo=True,
        ).order_by("nombre"),

        "Reacción": Ejercicio.objects.filter(
            categoria=Ejercicio.Categoria.REACCION,
            activo=True,
        ).order_by("nombre"),

        "Saque": Ejercicio.objects.filter(
            categoria=Ejercicio.Categoria.SAQUE,
            activo=True,
        ).order_by("nombre"),

        "Recepción": Ejercicio.objects.filter(
            categoria=Ejercicio.Categoria.RECEPCION,
            activo=True,
        ).order_by("nombre"),
    }

    contexto = {
        "entrenamiento": entrenamiento,
        "fecha": fecha,
        "hoy": timezone.localdate(),
        "turno": turno,
        "ayer": fecha - timedelta(days=1),
        "maniana": fecha + timedelta(days=1),
        "ejercicios_por_categoria": ejercicios_por_categoria,
        "ejercicios_guardados": ejercicios_guardados,
    }

    return render(
        request,
        "asistencia/cargar_ejercicios.html",
        contexto,
    )


@login_required
@require_POST
def guardar_ejercicios(request):
    entrenamiento_id = request.POST.get("entrenamiento_id")
    ejercicio_ids = request.POST.getlist("ejercicios")

    if not entrenamiento_id:
        messages.error(
            request,
            "No se pudo identificar el turno. Volvé a entrar desde la pantalla de asistencia.",
        )

        return redirect("inicio")

    entrenamiento = get_object_or_404(
        Entrenamiento,
        id=entrenamiento_id,
    )

    if turno_bloqueado(entrenamiento):
        return redirigir_turno_bloqueado(
            request,
            entrenamiento,
        )

    EjercicioTurno.objects.filter(
        entrenamiento=entrenamiento,
    ).delete()

    ejercicios_validos = (
        Ejercicio.objects
        .filter(
            id__in=ejercicio_ids,
            activo=True,
        )
    )

    for ejercicio in ejercicios_validos:
        EjercicioTurno.objects.create(
            entrenamiento=entrenamiento,
            ejercicio=ejercicio,
        )

    messages.success(
        request,
        f"Ejercicios guardados para el Turno {entrenamiento.turno}.",
    )

    return redirect(
        "dia_turno",
        fecha_str=entrenamiento.fecha.isoformat(),
        turno=entrenamiento.turno,
    )

@login_required
def seguimiento_semanal(request):
    jugadores = (
        Jugador.objects
        .filter(activo=True)
        .order_by("apellido", "nombre")
    )

    jugador_id = request.GET.get("jugador")
    fecha_str = request.GET.get("fecha")

    if fecha_str:
        try:
            fecha_base = datetime.strptime(
                fecha_str,
                "%Y-%m-%d",
            ).date()
        except ValueError:
            fecha_base = timezone.localdate()
    else:
        fecha_base = timezone.localdate()

    inicio_semana = fecha_base - timedelta(
        days=fecha_base.weekday()
    )

    fin_semana = inicio_semana + timedelta(days=4)

    dias_semana = [
        inicio_semana + timedelta(days=i)
        for i in range(5)
    ]

    semana_anterior = inicio_semana - timedelta(days=7)
    semana_siguiente = inicio_semana + timedelta(days=7)

    jugador_seleccionado = None
    filas = []
    resumen = None

    if jugador_id:
        jugador_seleccionado = (
            Jugador.objects
            .filter(
                id=jugador_id,
                activo=True,
            )
            .first()
        )

        if jugador_seleccionado:
            total_dias_programados = 0
            total_presentes = 0
            total_tardes = 0
            total_ausencias = 0
            total_dias_no_entrenados = 0
            total_ejercicios = 0
            total_trabajos = 0
            total_partidos = 0
            total_partidos_ganados = 0
            total_partidos_perdidos = 0

            for dia in dias_semana:
                entrenamientos_no_entrenados = (
                    Entrenamiento.objects
                    .filter(
                        fecha=dia,
                        no_se_entreno=True,
                    )
                    .order_by("turno")
                )

                turnos_no_entrenados = list(
                    entrenamientos_no_entrenados.values_list(
                        "turno",
                        flat=True,
                    )
                )

                motivos_no_entrenamiento = []

                for entrenamiento_no in entrenamientos_no_entrenados:
                    motivo_texto = (
                        entrenamiento_no.get_motivo_no_entrenamiento_display()
                    )

                    if entrenamiento_no.detalle_no_entrenamiento:
                        motivo_texto = (
                            f"{motivo_texto}: "
                            f"{entrenamiento_no.detalle_no_entrenamiento}"
                        )

                    motivos_no_entrenamiento.append({
                        "turno": entrenamiento_no.turno,
                        "motivo": motivo_texto,
                    })

                asistencias_dia = (
                    Asistencia.objects
                    .filter(
                        jugador=jugador_seleccionado,
                        entrenamiento__fecha=dia,
                    )
                    .select_related("entrenamiento")
                )

                turnos = list(
                    asistencias_dia.values_list(
                        "entrenamiento__turno",
                        flat=True,
                    )
                )

                estados = list(
                    asistencias_dia.values_list(
                        "estado",
                        flat=True,
                    )
                )

                if asistencias_dia.exists():
                    total_dias_programados += 1

                if turnos_no_entrenados:
                    total_dias_no_entrenados += 1

                entrenamientos_del_jugador_dia = [
                    asistencia.entrenamiento
                    for asistencia in asistencias_dia
                ]

                ejercicios_qs = (
                    EjercicioTurno.objects
                    .filter(
                        entrenamiento__in=entrenamientos_del_jugador_dia,
                    )
                    .select_related(
                        "ejercicio",
                        "entrenamiento",
                    )
                    .order_by(
                        "entrenamiento__turno",
                        "ejercicio__categoria",
                        "ejercicio__nombre",
                    )
                )

                ejercicios_por_turno = []
                turnos_dict = {}

                for item in ejercicios_qs:
                    turno_item = item.entrenamiento.turno
                    categoria = item.ejercicio.get_categoria_display()

                    if turno_item not in turnos_dict:
                        turnos_dict[turno_item] = {}

                    if categoria not in turnos_dict[turno_item]:
                        turnos_dict[turno_item][categoria] = []

                    turnos_dict[turno_item][categoria].append(
                        item.ejercicio.nombre
                    )

                for turno_item in sorted(turnos_dict.keys()):
                    ejercicios_por_turno.append({
                        "turno": turno_item,
                        "categorias": turnos_dict[turno_item],
                    })

                cantidad_ejercicios = ejercicios_qs.count()
                total_ejercicios += cantidad_ejercicios

                trabajos_qs = (
                    TrabajoTurno.objects
                    .filter(
                        entrenamiento__fecha=dia,
                    )
                    .filter(
                        Q(jugador_1=jugador_seleccionado)
                        | Q(jugador_2=jugador_seleccionado)
                    )
                    .select_related(
                        "entrenamiento",
                        "jugador_1",
                        "jugador_2",
                    )
                    .order_by(
                        "entrenamiento__turno",
                        "cambio",
                        "id",
                    )
                )

                trabajos_dia = []

                for trabajo in trabajos_qs:
                    if trabajo.tipo == TrabajoTurno.Tipo.PAREJA:
                        if trabajo.jugador_1_id == jugador_seleccionado.id:
                            companero = trabajo.jugador_2
                        else:
                            companero = trabajo.jugador_1

                        descripcion = f"Con {companero}"
                    else:
                        descripcion = trabajo.get_tipo_display()

                    trabajos_dia.append({
                        "turno": trabajo.entrenamiento.turno,
                        "cambio": trabajo.cambio,
                        "tipo": trabajo.tipo,
                        "tipo_texto": trabajo.get_tipo_display(),
                        "descripcion": descripcion,
                        "detalle": trabajo.detalle,
                    })

                cantidad_trabajos = len(trabajos_dia)
                total_trabajos += cantidad_trabajos

                partidos_qs = (
                    PartidoTurno.objects
                    .filter(
                        entrenamiento__fecha=dia,
                    )
                    .filter(
                        Q(jugador_1=jugador_seleccionado)
                        | Q(jugador_2=jugador_seleccionado)
                    )
                    .select_related(
                        "entrenamiento",
                        "jugador_1",
                        "jugador_2",
                    )
                    .prefetch_related("sets")
                    .order_by(
                        "entrenamiento__turno",
                        "id",
                    )
                )

                partidos_dia = []

                for partido in partidos_qs:
                    if partido.jugador_1_id == jugador_seleccionado.id:
                        rival = partido.jugador_2
                        sets_propios = partido.sets_jugador_1
                        sets_rival = partido.sets_jugador_2

                        sets_resultados = [
                            (
                                f"{set_partido.puntos_jugador_1}-"
                                f"{set_partido.puntos_jugador_2}"
                            )
                            for set_partido in partido.sets.all()
                        ]

                    else:
                        rival = partido.jugador_1
                        sets_propios = partido.sets_jugador_2
                        sets_rival = partido.sets_jugador_1

                        sets_resultados = [
                            (
                                f"{set_partido.puntos_jugador_2}-"
                                f"{set_partido.puntos_jugador_1}"
                            )
                            for set_partido in partido.sets.all()
                        ]

                    if partido.ganador == jugador_seleccionado:
                        resultado_clase = "success"
                        resultado_texto = "Victoria"
                        total_partidos_ganados += 1

                    elif partido.ganador:
                        resultado_clase = "danger"
                        resultado_texto = "Derrota"
                        total_partidos_perdidos += 1

                    else:
                        resultado_clase = "secondary"
                        resultado_texto = "Sin definir"

                    partidos_dia.append({
                        "turno": partido.entrenamiento.turno,
                        "rival": rival,
                        "resultado": f"{sets_propios}-{sets_rival}",
                        "resultado_clase": resultado_clase,
                        "resultado_texto": resultado_texto,
                        "sets": sets_resultados,
                        "detalle": partido.detalle,
                    })

                cantidad_partidos = len(partidos_dia)
                total_partidos += cantidad_partidos

                if "asistio" in estados:
                    asistencia_texto = "Asistió"
                    asistencia_clase = "success"
                    total_presentes += 1

                elif "tarde" in estados:
                    asistencia_texto = "Tarde"
                    asistencia_clase = "warning text-dark"
                    total_presentes += 1
                    total_tardes += 1

                elif "ausente" in estados:
                    asistencia_texto = "Ausente"
                    asistencia_clase = "danger"
                    total_ausencias += 1

                elif asistencias_dia.exists():
                    asistencia_texto = "Sin marcar"
                    asistencia_clase = "secondary"

                elif turnos_no_entrenados:
                    asistencia_texto = "No se entrenó"
                    asistencia_clase = "danger"

                elif (
                    cantidad_ejercicios > 0
                    or cantidad_trabajos > 0
                    or cantidad_partidos > 0
                ):
                    asistencia_texto = "Sin asistencia"
                    asistencia_clase = "secondary"

                else:
                    asistencia_texto = "Sin actividad"
                    asistencia_clase = "light text-dark"

                turnos_combinados = sorted(
                    set(turnos + turnos_no_entrenados)
                )

                filas.append({
                    "dia": dia,
                    "asistencia_texto": asistencia_texto,
                    "asistencia_clase": asistencia_clase,
                    "turnos": (
                        ", ".join(
                            str(turno_item)
                            for turno_item in turnos_combinados
                        )
                        if turnos_combinados
                        else "-"
                    ),
                    "no_se_entreno": bool(turnos_no_entrenados),
                    "turnos_no_entrenados": turnos_no_entrenados,
                    "motivos_no_entrenamiento": motivos_no_entrenamiento,
                    "ejercicios_por_turno": ejercicios_por_turno,
                    "cantidad_ejercicios": cantidad_ejercicios,
                    "trabajos_dia": trabajos_dia,
                    "cantidad_trabajos": cantidad_trabajos,
                    "partidos_dia": partidos_dia,
                    "cantidad_partidos": cantidad_partidos,
                })

            porcentaje = (
                round(
                    (
                        total_presentes
                        / total_dias_programados
                    ) * 100,
                    1,
                )
                if total_dias_programados
                else 0
            )

            resumen = {
                "total_dias_programados": total_dias_programados,
                "total_presentes": total_presentes,
                "total_tardes": total_tardes,
                "total_ausencias": total_ausencias,
                "total_dias_no_entrenados": total_dias_no_entrenados,
                "total_ejercicios": total_ejercicios,
                "total_trabajos": total_trabajos,
                "total_partidos": total_partidos,
                "total_partidos_ganados": total_partidos_ganados,
                "total_partidos_perdidos": total_partidos_perdidos,
                "porcentaje": porcentaje,
            }

    contexto = {
        "jugadores": jugadores,
        "jugador_seleccionado": jugador_seleccionado,
        "fecha_base": fecha_base,
        "inicio_semana": inicio_semana,
        "fin_semana": fin_semana,
        "semana_anterior": semana_anterior,
        "semana_siguiente": semana_siguiente,
        "filas": filas,
        "resumen": resumen,
    }

    return render(
        request,
        "asistencia/seguimiento_semanal.html",
        contexto,
    )


    

@login_required
def historial_jugador(request, jugador_id):
    jugador = get_object_or_404(
        Jugador,
        id=jugador_id,
    )

    hoy = timezone.localdate()
    fecha_desde_str = request.GET.get("desde")
    fecha_hasta_str = request.GET.get("hasta")

    if fecha_desde_str:
        try:
            fecha_desde = datetime.strptime(
                fecha_desde_str,
                "%Y-%m-%d",
            ).date()
        except ValueError:
            fecha_desde = hoy - timedelta(days=30)
    else:
        fecha_desde = hoy - timedelta(days=30)

    if fecha_hasta_str:
        try:
            fecha_hasta = datetime.strptime(
                fecha_hasta_str,
                "%Y-%m-%d",
            ).date()
        except ValueError:
            fecha_hasta = hoy
    else:
        fecha_hasta = hoy

    if fecha_desde > fecha_hasta:
        fecha_desde, fecha_hasta = fecha_hasta, fecha_desde

    inicio_mes = fecha_hasta.replace(day=1)

    asistencias = (
        Asistencia.objects
        .filter(
            jugador=jugador,
            entrenamiento__fecha__range=[
                fecha_desde,
                fecha_hasta,
            ],
            entrenamiento__no_se_entreno=False,
        )
        .select_related(
            "entrenamiento",
            "entrenamiento__entrenador",
            "entrenamiento__entrenador_responsable",
        )
        .order_by(
            "-entrenamiento__fecha",
            "entrenamiento__turno",
        )
    )

    total_asistencias = asistencias.count()

    total_presentes = asistencias.filter(
        Q(estado="asistio")
        | Q(estado="tarde")
    ).count()

    total_tardes = asistencias.filter(
        estado="tarde",
    ).count()

    total_ausencias = asistencias.filter(
        estado="ausente",
    ).count()

    total_pendientes = asistencias.filter(
        estado="pendiente",
    ).count()

    ausencias = asistencias.filter(
        estado="ausente",
    )

    ausencias_justificadas = (
        ausencias
        .exclude(motivo_ausencia="")
        .exclude(motivo_ausencia="sin_aviso")
        .count()
    )

    ausencias_sin_aviso = ausencias.filter(
        Q(motivo_ausencia="")
        | Q(motivo_ausencia="sin_aviso")
    ).count()

    porcentaje_asistencia = (
        round(
            (
                total_presentes
                / total_asistencias
            ) * 100,
            1,
        )
        if total_asistencias
        else 0
    )

    asistencias_mes = asistencias.filter(
        entrenamiento__fecha__range=[
            inicio_mes,
            fecha_hasta,
        ]
    )

    mes_total = asistencias_mes.count()

    mes_presentes = asistencias_mes.filter(
        Q(estado="asistio")
        | Q(estado="tarde")
    ).count()

    mes_ausencias = asistencias_mes.filter(
        estado="ausente",
    ).count()

    mes_tardes = asistencias_mes.filter(
        estado="tarde",
    ).count()

    mes_porcentaje = (
        round(
            (
                mes_presentes
                / mes_total
            ) * 100,
            1,
        )
        if mes_total
        else 0
    )

    for asistencia in asistencias:
        asistencia.ejercicios_turno = (
            EjercicioTurno.objects
            .filter(
                entrenamiento=asistencia.entrenamiento,
            )
            .select_related("ejercicio")
            .order_by(
                "ejercicio__categoria",
                "ejercicio__nombre",
            )
        )

    entrenamientos_del_jugador = [
        asistencia.entrenamiento
        for asistencia in asistencias
    ]

    ejercicios_turno = (
        EjercicioTurno.objects
        .filter(
            entrenamiento__in=entrenamientos_del_jugador,
        )
        .select_related(
            "entrenamiento",
            "ejercicio",
        )
        .order_by(
            "-entrenamiento__fecha",
            "entrenamiento__turno",
            "ejercicio__categoria",
            "ejercicio__nombre",
        )
    )

    ejercicios = []

    for item in ejercicios_turno:
        ejercicios.append({
            "fecha": item.entrenamiento.fecha,
            "turno": item.entrenamiento.turno,
            "ejercicio": item.ejercicio,
        })

    ejercicios_frecuentes = (
        ejercicios_turno
        .values(
            "ejercicio__nombre",
            "ejercicio__categoria",
        )
        .annotate(total=Count("id"))
        .order_by("-total")[:8]
    )

    partidos = (
        PartidoTurno.objects
        .filter(
            entrenamiento__fecha__range=[
                fecha_desde,
                fecha_hasta,
            ],
        )
        .filter(
            Q(jugador_1=jugador)
            | Q(jugador_2=jugador)
        )
        .select_related(
            "entrenamiento",
            "jugador_1",
            "jugador_2",
            "ganador",
        )
        .prefetch_related("sets")
        .order_by(
            "-entrenamiento__fecha",
            "entrenamiento__turno",
            "-id",
        )
    )

    total_partidos = partidos.count()

    total_partidos_ganados = partidos.filter(
        ganador=jugador,
    ).count()

    total_partidos_perdidos = partidos.exclude(
        ganador__isnull=True,
    ).exclude(
        ganador=jugador,
    ).count()

    porcentaje_victorias = (
        round(
            (
                total_partidos_ganados
                / total_partidos
            ) * 100,
            1,
        )
        if total_partidos
        else 0
    )

    trabajos = (
        TrabajoTurno.objects
        .filter(
            entrenamiento__fecha__range=[
                fecha_desde,
                fecha_hasta,
            ],
        )
        .filter(
            Q(jugador_1=jugador)
            | Q(jugador_2=jugador)
        )
        .select_related(
            "entrenamiento",
            "jugador_1",
            "jugador_2",
        )
        .order_by(
            "-entrenamiento__fecha",
            "entrenamiento__turno",
            "cambio",
        )
    )

    observaciones = (
        ObservacionJugador.objects
        .filter(
            jugador=jugador,
            entrenamiento__fecha__range=[
                fecha_desde,
                fecha_hasta,
            ],
        )
        .select_related(
            "entrenamiento",
            "creada_por",
        )
        .order_by(
            "-entrenamiento__fecha",
            "-creada_el",
        )
    )

    turnos_no_entrenados = (
        Entrenamiento.objects
        .filter(
            fecha__range=[
                fecha_desde,
                fecha_hasta,
            ],
            no_se_entreno=True,
        )
        .order_by(
            "-fecha",
            "turno",
        )
    )

    resumen = {
        "total_asistencias": total_asistencias,
        "total_presentes": total_presentes,
        "total_tardes": total_tardes,
        "total_ausencias": total_ausencias,
        "total_pendientes": total_pendientes,
        "ausencias_justificadas": ausencias_justificadas,
        "ausencias_sin_aviso": ausencias_sin_aviso,
        "porcentaje_asistencia": porcentaje_asistencia,

        "mes_total": mes_total,
        "mes_presentes": mes_presentes,
        "mes_ausencias": mes_ausencias,
        "mes_tardes": mes_tardes,
        "mes_porcentaje": mes_porcentaje,

        "total_partidos": total_partidos,
        "total_partidos_ganados": total_partidos_ganados,
        "total_partidos_perdidos": total_partidos_perdidos,
        "porcentaje_victorias": porcentaje_victorias,

        "total_trabajos": trabajos.count(),
        "total_ejercicios": len(ejercicios),
        "total_observaciones": observaciones.count(),
        "turnos_no_entrenados": turnos_no_entrenados.count(),
    }

    contexto = {
        "jugador": jugador,
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
        "inicio_mes": inicio_mes,
        "asistencias": asistencias,
        "partidos": partidos,
        "trabajos": trabajos,
        "ejercicios": ejercicios,
        "ejercicios_frecuentes": ejercicios_frecuentes,
        "observaciones": observaciones,
        "turnos_no_entrenados": turnos_no_entrenados,
        "resumen": resumen,

        # Compatibilidad con el template viejo, por si alguna parte todavía los usa.
        "total_asistencias": total_asistencias,
        "total_presentes": total_presentes,
        "total_tardes": total_tardes,
        "total_ausencias": total_ausencias,
        "porcentaje_asistencia": porcentaje_asistencia,
        "total_partidos": total_partidos,
        "total_partidos_ganados": total_partidos_ganados,
        "total_partidos_perdidos": total_partidos_perdidos,
    }

    return render(
        request,
        "asistencia/historial_jugador.html",
        contexto,
    )


@login_required
def reportes(request):
    hoy = timezone.localdate()
    inicio_semana = hoy - timedelta(days=hoy.weekday())
    inicio_mes = hoy.replace(day=1)

    turnos_no_entrenados_semana = (
        Entrenamiento.objects
        .filter(
            fecha__range=[
                inicio_semana,
                hoy,
            ],
            no_se_entreno=True,
        )
        .order_by(
            "-fecha",
            "turno",
        )
    )

    turnos_no_entrenados_mes = (
        Entrenamiento.objects
        .filter(
            fecha__range=[
                inicio_mes,
                hoy,
            ],
            no_se_entreno=True,
        )
        .order_by(
            "-fecha",
            "turno",
        )
    )

    total_turnos_no_entrenados_semana = (
        turnos_no_entrenados_semana.count()
    )

    total_turnos_no_entrenados_mes = (
        turnos_no_entrenados_mes.count()
    )

    dias_no_entrenados_semana = (
        turnos_no_entrenados_semana
        .values("fecha")
        .distinct()
        .count()
    )

    dias_no_entrenados_mes = (
        turnos_no_entrenados_mes
        .values("fecha")
        .distinct()
        .count()
    )

    motivos_no_entrenamiento_mes = (
        turnos_no_entrenados_mes
        .values("motivo_no_entrenamiento")
        .annotate(total=Count("id"))
        .order_by("-total")
    )

    motivos_no_entrenamiento = []

    opciones_motivos = dict(
        Entrenamiento.MotivoNoEntrenamiento.choices
    )

    for item in motivos_no_entrenamiento_mes:
        motivo_codigo = item["motivo_no_entrenamiento"]

        if motivo_codigo:
            motivo_texto = opciones_motivos.get(
                motivo_codigo,
                motivo_codigo,
            )
        else:
            motivo_texto = "Sin motivo"

        motivos_no_entrenamiento.append({
            "motivo": motivo_texto,
            "total": item["total"],
        })

    datos = []

    total_jugadores = 0
    total_semana_presentes = 0
    total_semana_ausentes = 0
    total_mes_presentes = 0
    total_mes_ausentes = 0

    for jugador in Jugador.objects.filter(activo=True).order_by("apellido", "nombre"):
        total_jugadores += 1

        asistencias_semana = Asistencia.objects.filter(
            jugador=jugador,
            entrenamiento__fecha__range=[
                inicio_semana,
                hoy,
            ],
            entrenamiento__no_se_entreno=False,
        )

        asistencias_mes = Asistencia.objects.filter(
            jugador=jugador,
            entrenamiento__fecha__range=[
                inicio_mes,
                hoy,
            ],
            entrenamiento__no_se_entreno=False,
        )

        semana_total = asistencias_semana.count()

        semana_presentes = asistencias_semana.filter(
            Q(estado="asistio")
            | Q(estado="tarde")
        ).count()

        semana_tardes = asistencias_semana.filter(
            estado="tarde"
        ).count()

        semana_ausentes = asistencias_semana.filter(
            estado="ausente"
        ).count()

        semana_pendientes = asistencias_semana.filter(
            estado="pendiente"
        ).count()

        semana_porcentaje = (
            round(
                (
                    semana_presentes
                    / semana_total
                ) * 100,
                1,
            )
            if semana_total
            else 0
        )

        mes_total = asistencias_mes.count()

        mes_presentes = asistencias_mes.filter(
            Q(estado="asistio")
            | Q(estado="tarde")
        ).count()

        mes_tardes = asistencias_mes.filter(
            estado="tarde"
        ).count()

        ausencias_mes = asistencias_mes.filter(
            estado="ausente"
        )

        mes_ausentes = ausencias_mes.count()

        mes_pendientes = asistencias_mes.filter(
            estado="pendiente"
        ).count()

        mes_ausencias_justificadas = (
            ausencias_mes
            .exclude(motivo_ausencia="")
            .exclude(motivo_ausencia="sin_aviso")
            .count()
        )

        mes_ausencias_sin_aviso = ausencias_mes.filter(
            Q(motivo_ausencia="sin_aviso")
            | Q(motivo_ausencia="")
        ).count()

        motivo_mas_frecuente = (
            ausencias_mes
            .exclude(motivo_ausencia="")
            .values("motivo_ausencia")
            .annotate(total=Count("id"))
            .order_by("-total")
            .first()
        )

        if motivo_mas_frecuente:
            motivo_codigo = motivo_mas_frecuente[
                "motivo_ausencia"
            ]

            motivo_mas_frecuente_texto = dict(
                Asistencia.MotivoAusencia.choices
            ).get(
                motivo_codigo,
                motivo_codigo,
            )
        else:
            motivo_mas_frecuente_texto = "-"

        mes_porcentaje = (
            round(
                (
                    mes_presentes
                    / mes_total
                ) * 100,
                1,
            )
            if mes_total
            else 0
        )

        if mes_total == 0:
            estado_clase = "secondary"
            estado_texto = "Sin datos"

        elif mes_porcentaje >= 80:
            estado_clase = "success"
            estado_texto = "Buena asistencia"

        elif mes_porcentaje >= 50:
            estado_clase = "warning text-dark"
            estado_texto = "Asistencia media"

        else:
            estado_clase = "danger"
            estado_texto = "Baja asistencia"

        total_semana_presentes += semana_presentes
        total_semana_ausentes += semana_ausentes
        total_mes_presentes += mes_presentes
        total_mes_ausentes += mes_ausentes

        datos.append({
            "jugador": jugador,

            "semana_total": semana_total,
            "semana_presentes": semana_presentes,
            "semana_tardes": semana_tardes,
            "semana_ausentes": semana_ausentes,
            "semana_pendientes": semana_pendientes,
            "semana_porcentaje": semana_porcentaje,

            "mes_total": mes_total,
            "mes_presentes": mes_presentes,
            "mes_tardes": mes_tardes,
            "mes_ausentes": mes_ausentes,
            "mes_pendientes": mes_pendientes,
            "mes_porcentaje": mes_porcentaje,

            "mes_ausencias_justificadas": (
                mes_ausencias_justificadas
            ),
            "mes_ausencias_sin_aviso": (
                mes_ausencias_sin_aviso
            ),
            "motivo_mas_frecuente": (
                motivo_mas_frecuente_texto
            ),

            "estado_clase": estado_clase,
            "estado_texto": estado_texto,
        })

    resumen_general = {
        "total_jugadores": total_jugadores,
        "total_semana_presentes": total_semana_presentes,
        "total_semana_ausentes": total_semana_ausentes,
        "total_mes_presentes": total_mes_presentes,
        "total_mes_ausentes": total_mes_ausentes,
        "dias_no_entrenados_semana": dias_no_entrenados_semana,
        "dias_no_entrenados_mes": dias_no_entrenados_mes,
        "total_turnos_no_entrenados_semana": total_turnos_no_entrenados_semana,
        "total_turnos_no_entrenados_mes": total_turnos_no_entrenados_mes,
    }

    return render(
        request,
        "asistencia/reportes.html",
        {
            "datos": datos,
            "hoy": hoy,
            "inicio_semana": inicio_semana,
            "inicio_mes": inicio_mes,
            "resumen_general": resumen_general,
            "turnos_no_entrenados_semana": turnos_no_entrenados_semana,
            "turnos_no_entrenados_mes": turnos_no_entrenados_mes,
            "motivos_no_entrenamiento": motivos_no_entrenamiento,
        },
    )

@login_required
def exportar_reporte_mensual(request):
    hoy = timezone.localdate()
    fecha_str = request.GET.get("fecha")

    if fecha_str:
        try:
            fecha_base = datetime.strptime(
                fecha_str,
                "%Y-%m-%d",
            ).date()
        except ValueError:
            fecha_base = hoy
    else:
        fecha_base = hoy

    inicio_mes = fecha_base.replace(day=1)

    if inicio_mes.month == 12:
        inicio_mes_siguiente = inicio_mes.replace(
            year=inicio_mes.year + 1,
            month=1,
        )
    else:
        inicio_mes_siguiente = inicio_mes.replace(
            month=inicio_mes.month + 1,
        )

    fin_mes = inicio_mes_siguiente - timedelta(days=1)

    wb = Workbook()

    # =========================
    # ESTILOS
    # =========================
    titulo_font = Font(bold=True, size=14, color="FFFFFF")
    seccion_font = Font(bold=True, size=12, color="FFFFFF")
    header_font = Font(bold=True, color="FFFFFF")
    normal_bold = Font(bold=True)

    titulo_fill = PatternFill(
        start_color="111827",
        end_color="111827",
        fill_type="solid",
    )

    seccion_fill = PatternFill(
        start_color="1F2937",
        end_color="1F2937",
        fill_type="solid",
    )

    header_fill = PatternFill(
        start_color="374151",
        end_color="374151",
        fill_type="solid",
    )

    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def ajustar_columnas(ws):
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = left

        for column_index in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(column_index)

            for row_index in range(1, ws.max_row + 1):
                cell = ws.cell(
                    row=row_index,
                    column=column_index,
                )

                value = cell.value

                if value is not None:
                    max_length = max(
                        max_length,
                        len(str(value)),
                    )

            ws.column_dimensions[column_letter].width = min(
                max_length + 3,
                38,
            )

    def aplicar_header(ws, row_number):
        for cell in ws[row_number]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

    def aplicar_titulo(ws, row_number, texto):
        ws.merge_cells(
            start_row=row_number,
            start_column=1,
            end_row=row_number,
            end_column=8,
        )

        cell = ws.cell(row=row_number, column=1)
        cell.value = texto
        cell.font = titulo_font
        cell.fill = titulo_fill
        cell.alignment = center

    def aplicar_seccion(ws, row_number, texto):
        ws.merge_cells(
            start_row=row_number,
            start_column=1,
            end_row=row_number,
            end_column=8,
        )

        cell = ws.cell(row=row_number, column=1)
        cell.value = texto
        cell.font = seccion_font
        cell.fill = seccion_fill
        cell.alignment = left

    # =========================
    # DATOS BASE
    # =========================
    asistencias_mes = Asistencia.objects.filter(
        entrenamiento__fecha__range=[
            inicio_mes,
            fin_mes,
        ],
        entrenamiento__no_se_entreno=False,
    ).select_related(
        "jugador",
        "entrenamiento",
        "entrenamiento__entrenador",
        "entrenamiento__entrenador_responsable",
    )

    turnos_no_entrenados = Entrenamiento.objects.filter(
        fecha__range=[
            inicio_mes,
            fin_mes,
        ],
        no_se_entreno=True,
    ).order_by(
        "fecha",
        "turno",
    )

    partidos_mes = (
        PartidoTurno.objects
        .filter(
            entrenamiento__fecha__range=[
                inicio_mes,
                fin_mes,
            ],
        )
        .select_related(
            "entrenamiento",
            "jugador_1",
            "jugador_2",
        )
        .prefetch_related("sets")
        .order_by(
            "entrenamiento__fecha",
            "entrenamiento__turno",
            "id",
        )
    )

    jugadores = Jugador.objects.all().order_by(
        "apellido",
        "nombre",
    )

    jugadores_activos = jugadores.filter(
        activo=True,
    ).count()

    jugadores_totales = jugadores.count()

    total_presentes = asistencias_mes.filter(
        Q(estado="asistio")
        | Q(estado="tarde")
    ).count()

    total_tardes = asistencias_mes.filter(
        estado="tarde",
    ).count()

    total_ausencias = asistencias_mes.filter(
        estado="ausente",
    ).count()

    total_pendientes = asistencias_mes.filter(
        estado="pendiente",
    ).count()

    # =========================
    # HOJA 1: PANEL GENERAL
    # =========================
    ws = wb.active
    ws.title = "Panel general"

    fila = 1

    aplicar_titulo(
        ws,
        fila,
        f"REPORTE MENSUAL - {inicio_mes.strftime('%m/%Y')}",
    )

    fila += 2

    aplicar_seccion(ws, fila, "RESUMEN DEL MES")
    fila += 1

    ws.append(["Campo", "Valor"])
    aplicar_header(ws, fila)
    fila += 1

    resumen_filas = [
        ["Mes", inicio_mes.strftime("%m/%Y")],
        ["Desde", inicio_mes.strftime("%d/%m/%Y")],
        ["Hasta", fin_mes.strftime("%d/%m/%Y")],
        ["Jugadores totales", jugadores_totales],
        ["Jugadores activos", jugadores_activos],
        ["Presentes", total_presentes],
        ["Tardes", total_tardes],
        ["Ausencias", total_ausencias],
        ["Pendientes", total_pendientes],
        ["Turnos sin entrenamiento", turnos_no_entrenados.count()],
        ["Partidos cargados", partidos_mes.count()],
    ]

    for item in resumen_filas:
        ws.append(item)
        fila += 1

    fila += 2

    aplicar_seccion(ws, fila, "JUGADORES")
    fila += 1

    ws.append([
        "ID",
        "Nombre",
        "Apellido",
        "Jugador completo",
        "Activo",
    ])

    aplicar_header(ws, fila)
    fila += 1

    for jugador in jugadores:
        ws.append([
            jugador.id,
            jugador.nombre,
            jugador.apellido,
            str(jugador),
            "Sí" if jugador.activo else "No",
        ])

        fila += 1

    fila += 2

    aplicar_seccion(ws, fila, "ASISTENCIA POR JUGADOR")
    fila += 1

    ws.append([
        "Jugador",
        "Turnos cargados",
        "Presentes",
        "Tardes",
        "Ausencias",
        "Pendientes",
        "% asistencia",
        "Motivo frecuente",
    ])

    aplicar_header(ws, fila)
    fila += 1

    datos_asistencia_jugadores = []

    for jugador in jugadores.filter(activo=True):
        asistencias_jugador = asistencias_mes.filter(
            jugador=jugador,
        )

        total = asistencias_jugador.count()

        presentes = asistencias_jugador.filter(
            Q(estado="asistio")
            | Q(estado="tarde")
        ).count()

        tardes = asistencias_jugador.filter(
            estado="tarde",
        ).count()

        ausencias = asistencias_jugador.filter(
            estado="ausente",
        )

        ausentes = ausencias.count()

        pendientes = asistencias_jugador.filter(
            estado="pendiente",
        ).count()

        porcentaje = (
            round(
                (
                    presentes
                    / total
                ) * 100,
                1,
            )
            if total
            else 0
        )

        justificadas = (
            ausencias
            .exclude(motivo_ausencia="")
            .exclude(motivo_ausencia="sin_aviso")
            .count()
        )

        sin_aviso = ausencias.filter(
            Q(motivo_ausencia="")
            | Q(motivo_ausencia="sin_aviso")
        ).count()

        motivo_mas_frecuente = (
            ausencias
            .exclude(motivo_ausencia="")
            .values("motivo_ausencia")
            .annotate(total=Count("id"))
            .order_by("-total")
            .first()
        )

        if motivo_mas_frecuente:
            motivo_codigo = motivo_mas_frecuente["motivo_ausencia"]

            motivo_texto = dict(
                Asistencia.MotivoAusencia.choices
            ).get(
                motivo_codigo,
                motivo_codigo,
            )
        else:
            motivo_texto = "-"

        datos_asistencia_jugadores.append({
            "jugador": jugador,
            "total": total,
            "presentes": presentes,
            "tardes": tardes,
            "ausentes": ausentes,
            "pendientes": pendientes,
            "porcentaje": porcentaje,
            "justificadas": justificadas,
            "sin_aviso": sin_aviso,
            "motivo_texto": motivo_texto,
        })

        ws.append([
            str(jugador),
            total,
            presentes,
            tardes,
            ausentes,
            pendientes,
            porcentaje,
            motivo_texto,
        ])

        fila += 1

    fila += 2

    aplicar_seccion(ws, fila, "TURNOS SIN ENTRENAMIENTO")
    fila += 1

    ws.append([
        "Fecha",
        "Turno",
        "Motivo",
        "Detalle",
    ])

    aplicar_header(ws, fila)
    fila += 1

    if turnos_no_entrenados.exists():
        for entrenamiento in turnos_no_entrenados:
            ws.append([
                entrenamiento.fecha.strftime("%d/%m/%Y"),
                entrenamiento.turno,
                entrenamiento.get_motivo_no_entrenamiento_display(),
                entrenamiento.detalle_no_entrenamiento or "-",
            ])

            fila += 1
    else:
        ws.append([
            "-",
            "-",
            "No hubo turnos sin entrenamiento",
            "-",
        ])

        fila += 1

    ajustar_columnas(ws)

    # =========================
    # HOJA 2: ASISTENCIA DETALLADA
    # =========================
    ws = wb.create_sheet("Asistencia detallada")

    ws.append([
        "Fecha",
        "Turno",
        "Jugador",
        "Estado",
        "Motivo ausencia",
        "Detalle ausencia",
        "Entrenador",
    ])

    aplicar_header(ws, 1)

    asistencias_detalladas = asistencias_mes.order_by(
        "entrenamiento__fecha",
        "entrenamiento__turno",
        "jugador__apellido",
        "jugador__nombre",
    )

    for asistencia in asistencias_detalladas:
        entrenamiento = asistencia.entrenamiento

        if entrenamiento.entrenador_responsable:
            entrenador = str(entrenamiento.entrenador_responsable)
        elif entrenamiento.entrenador:
            entrenador = (
                entrenamiento.entrenador.get_full_name()
                or entrenamiento.entrenador.username
            )
        else:
            entrenador = "Sin entrenador"

        ws.append([
            entrenamiento.fecha.strftime("%d/%m/%Y"),
            entrenamiento.turno,
            str(asistencia.jugador),
            asistencia.get_estado_display(),
            asistencia.get_motivo_ausencia_display()
            if asistencia.motivo_ausencia
            else "-",
            asistencia.detalle_ausencia or "-",
            entrenador,
        ])

    ajustar_columnas(ws)

    # =========================
    # HOJA 3: AUSENCIAS
    # =========================
    ws = wb.create_sheet("Ausencias")

    ws.append([
        "Fecha",
        "Turno",
        "Jugador",
        "Motivo",
        "Detalle",
        "Entrenador",
    ])

    aplicar_header(ws, 1)

    ausencias_mes = asistencias_mes.filter(
        estado="ausente",
    ).order_by(
        "entrenamiento__fecha",
        "entrenamiento__turno",
        "jugador__apellido",
        "jugador__nombre",
    )

    for asistencia in ausencias_mes:
        entrenamiento = asistencia.entrenamiento

        if entrenamiento.entrenador_responsable:
            entrenador = str(entrenamiento.entrenador_responsable)
        elif entrenamiento.entrenador:
            entrenador = (
                entrenamiento.entrenador.get_full_name()
                or entrenamiento.entrenador.username
            )
        else:
            entrenador = "Sin entrenador"

        ws.append([
            entrenamiento.fecha.strftime("%d/%m/%Y"),
            entrenamiento.turno,
            str(asistencia.jugador),
            asistencia.get_motivo_ausencia_display()
            if asistencia.motivo_ausencia
            else "Sin aviso",
            asistencia.detalle_ausencia or "-",
            entrenador,
        ])

    ajustar_columnas(ws)

    # =========================
    # HOJA 4: PARTIDOS
    # =========================
    ws = wb.create_sheet("Partidos")

    ws.append([
        "Fecha",
        "Turno",
        "Jugador 1",
        "Jugador 2",
        "Resultado",
        "Ganador",
        "Sets",
        "Detalle",
    ])

    aplicar_header(ws, 1)

    for partido in partidos_mes:
        sets_jugador_1 = partido.sets_jugador_1
        sets_jugador_2 = partido.sets_jugador_2

        if sets_jugador_1 > sets_jugador_2:
            ganador = str(partido.jugador_1)
        elif sets_jugador_2 > sets_jugador_1:
            ganador = str(partido.jugador_2)
        else:
            ganador = "-"

        sets_texto = []

        for set_partido in partido.sets.all():
            sets_texto.append(
                f"{set_partido.puntos_jugador_1}-{set_partido.puntos_jugador_2}"
            )

        ws.append([
            partido.entrenamiento.fecha.strftime("%d/%m/%Y"),
            partido.entrenamiento.turno,
            str(partido.jugador_1),
            str(partido.jugador_2),
            f"{sets_jugador_1}-{sets_jugador_2}",
            ganador,
            " / ".join(sets_texto) if sets_texto else "-",
            partido.detalle or "-",
        ])

    ajustar_columnas(ws)

    # =========================
    # RESPUESTA DESCARGA
    # =========================
    nombre_archivo = (
        f"reporte_mensual_"
        f"{inicio_mes.strftime('%Y_%m')}.xlsx"
    )

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    response["Content-Disposition"] = (
        f'attachment; filename="{nombre_archivo}"'
    )

    wb.save(response)

    return response

@login_required
def dashboard_mensual(request):
    hoy = timezone.localdate()

    contexto = {
        "fecha_base": hoy,
        "inicio_mes": hoy.replace(day=1),
        "fin_mes": hoy,
        "mes_anterior": hoy,
        "mes_siguiente": hoy,
        "resumen": {
            "jugadores_activos": Jugador.objects.filter(activo=True).count(),
            "turnos_realizados": 0,
            "turnos_no_entrenados": 0,
            "dias_no_entrenados": 0,
            "total_presentes": 0,
            "total_ausentes": 0,
            "total_tardes": 0,
            "total_partidos": 0,
            "motivo_principal": "-",
        },
        "motivos": [],
        "turnos_no_entrenados": [],
        "ranking_asistencia": [],
        "ranking_ausencias": [],
        "ranking_partidos": [],
        "mejor_asistencia": None,
        "mas_ausencias": None,
        "mas_partidos": None,
    }

    return render(
        request,
        "asistencia/dashboard_mensual.html",
        contexto,
    )


@login_required
def acerca(request):
    return render(request, "asistencia/acerca.html")

@login_required
def perfil(request):
    if request.method == "POST":
        form = PerfilForm(request.POST, instance=request.user)

        if form.is_valid():
            form.save()
            messages.success(request, "Perfil actualizado correctamente.")
            return redirect("perfil")
    else:
        form = PerfilForm(instance=request.user)

    return render(request, "asistencia/perfil.html", {
        "form": form,
    })
    
@login_required
@require_POST
def eliminar_observacion_jugador(request, observacion_id):
    observacion = get_object_or_404(
        ObservacionJugador,
        id=observacion_id,
    )

    entrenamiento = observacion.entrenamiento
    jugador = observacion.jugador

    if turno_bloqueado(entrenamiento):
        return redirigir_turno_bloqueado(
            request,
            entrenamiento,
        )

    observacion.delete()

    messages.success(
        request,
        f"Observación eliminada de {jugador}.",
    )

    return redirect_dia_turno(
        entrenamiento,
        request.POST.get("volver_a", ""),
    )
    
@login_required
def editar_observacion_jugador(request, observacion_id):
    observacion = get_object_or_404(
        ObservacionJugador,
        id=observacion_id,
    )

    entrenamiento = observacion.entrenamiento

    if request.method == "POST" and turno_bloqueado(entrenamiento):
        return redirigir_turno_bloqueado(
            request,
            entrenamiento,
        )

    if request.method == "POST":
        form = ObservacionJugadorForm(
            request.POST,
            instance=observacion,
        )

        if form.is_valid():
            form.save()

            messages.success(
                request,
                "Observación actualizada correctamente.",
            )

            return redirect(
                "dia_turno",
                fecha_str=entrenamiento.fecha.isoformat(),
                turno=entrenamiento.turno,
            )
    else:
        form = ObservacionJugadorForm(
            instance=observacion,
        )

    return render(
        request,
        "asistencia/editar_observacion_jugador.html",
        {
            "form": form,
            "observacion": observacion,
            "entrenamiento": entrenamiento,
        },
    )
    
@login_required
def editar_partido_turno(request, partido_id):
    partido = get_object_or_404(
        PartidoTurno.objects.select_related("entrenamiento"),
        id=partido_id,
    )

    entrenamiento = partido.entrenamiento

    if request.method == "POST" and turno_bloqueado(entrenamiento):
        return redirigir_turno_bloqueado(
            request,
            entrenamiento,
        )

    if request.method == "POST":
        partido_form = PartidoTurnoForm(
            request.POST,
            instance=partido,
            entrenamiento=entrenamiento,
        )

        sets_formset = SetPartidoFormSet(
            request.POST,
            instance=partido,
            prefix="sets",
        )

        if partido_form.is_valid() and sets_formset.is_valid():
            with transaction.atomic():
                partido = partido_form.save()
                sets = sets_formset.save(commit=False)

                numero_set = 1

                for set_partido in sets:
                    set_partido.partido = partido
                    set_partido.numero = numero_set
                    set_partido.save()
                    numero_set += 1

                for set_eliminado in sets_formset.deleted_objects:
                    set_eliminado.delete()

            messages.success(
                request,
                "El partido se actualizó correctamente.",
            )

            return redirect(
                "dia_turno",
                fecha_str=entrenamiento.fecha.isoformat(),
                turno=entrenamiento.turno,
            )

    else:
        partido_form = PartidoTurnoForm(
            instance=partido,
            entrenamiento=entrenamiento,
        )

        sets_formset = SetPartidoFormSet(
            instance=partido,
            prefix="sets",
        )

    return render(
        request,
        "asistencia/editar_partido_turno.html",
        {
            "partido": partido,
            "entrenamiento": entrenamiento,
            "partido_form": partido_form,
            "sets_formset": sets_formset,
        },
    )


@login_required
@require_POST
def eliminar_partido_turno(request, partido_id):
    partido = get_object_or_404(
        PartidoTurno.objects.select_related("entrenamiento"),
        id=partido_id,
    )

    entrenamiento = partido.entrenamiento

    if turno_bloqueado(entrenamiento):
        return redirigir_turno_bloqueado(
            request,
            entrenamiento,
        )

    partido.delete()

    messages.success(
        request,
        "El partido se eliminó correctamente.",
    )

    return redirect_dia_turno(
        entrenamiento,
        request.POST.get("volver_a", ""),
    )
    